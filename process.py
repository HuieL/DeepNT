
############## velocity计算-old

import sys
import os
import pandas as pd
from matplotlib import pyplot as plt
import math
# import xlwt
# import xlsxwriter

# plt.rcParams['font.sans-serif']=['SimHei']#用来正常显示中文标签
# plt.rcParams['axes.unicode_minus']=False#用来正常显示负号

# excel_path = 'Anaheim_Network.xlsx'
# # filenames = os.listdir(path)
# # f = xlwt.Workbook(encoding='utf-8', style_compression=0)  # 新建一个excel
# # sheet = f.add_sheet('sheet1')  # 新建一个sheet

# # for i in filenames:
# #     excel_path = r'file name/' + i
# f = open(excel_path,'rb')
# #     print(f)
# df = pd.read_excel(f)
# print(df)
# #定义新列
# col_name = df.columns.tolist()
# col_name.insert(10,'velocity') #设置新增列的位置和名称
# wb = df.reindex(columns = col_name)
# df['velocity'] = df['length'] / df['free_flow_time'] #计算方式，根据自己设定

# #To save it back as Excel
# df.to_excel('new.xlsx') #Write DateFrame back as Excel file



############## velocity计算-new0926
'''

import openpyxl
 
workbook=openpyxl.load_workbook("Anaheim_Network.xlsx")
worksheet=workbook.worksheets[0]
 
#在第一列之前插入一列
# worksheet.insert_cols(4)  #
 
for index,col in enumerate(worksheet.rows):
    print(index)
    print(col)
    if index==0:
        continue
    else:
        col[4].value=col[2].value/col[3].value
        col[0].value=col[0].value-1
        col[1].value=col[1].value-1
#枚举出来是tuple类型，从0开始计数
 
workbook.save(filename="new.xlsx")

'''

##################构成图
#例子
'''
class graphMatrix:
    def __init__(self):
        self.dict={}  # 结点对应矩阵下标的字典
        self.length=0 # 矩阵的长度
        self.matrix=[]

    # 添加节点
    def addVertex(self,key):
        self.dict[key]=self.length # 键(顶点名称)->值(在矩阵中的下标)
        self.dict[self.length]=key # 键(在矩阵中的下标)->值(顶点名称)


        lst=[]
        for i in range(self.length):  # 创建一行长度为self.length,全是0的列表
            lst.append(0)
        self.matrix.append(lst)

        for row in self.matrix:  # 为矩阵中每行添加新的一个单元，单元中值为0
            row.append(0)

        self.length += 1

    # 添加无向边
    def addNoDirectLine(self,start,ends): #起始字符，终点(字符+权值)数组
        startIndex=self.dict[start] # 起点在二维矩阵中的下标
        for row in ends:  # 遍历终点数组
            endKey=row[0]
            endIndex=self.dict[endKey]  # 终点在二维矩阵中的下标
            weight=row[1]
            self.matrix[startIndex][endIndex]=weight  # 权值赋值
            self.matrix[endIndex][startIndex]=weight

    # 添加有向边
    def addDirectLine(self,start,ends):
        startIndex=self.dict[start] # 起点在二维矩阵中的下标
        for row in ends:  # 遍历终点数组
            endKey=row[0]
            endIndex=self.dict[endKey]  # 终点在二维矩阵中的下标
            weight=row[1]
            self.matrix[startIndex][endIndex]=weight  # 权值赋值

def main():

    lst=['北京','天津','郑州','青岛']
    ends = [[['天津',138],['郑州',689]],[['郑州',700],['青岛',597]],[['青岛',725]]]
    gm=graphMatrix()
    for k in lst:
        gm.addVertex(k)
    for i in range(len(ends)):
        gm.addNoDirectLine(lst[i],ends[i])
    
    

    # print(gm.dict)
    # print(gm.matrix)
    for i in range(len(ends)):
        gm.addNoDirectLine(lst[i],ends[i])

    print('\n')
    for i in range(len(gm.matrix)):
        # if i==0:
        #     print('    ',end='')
        #     for k in lst:
        #         print(k,end=' ')
        #     print('')
        # print(gm.dict[i],gm.matrix[i])
        print(gm.matrix[i])

if __name__ == '__main__':
    main()
'''



'''



import sys
import os
import pandas as pd
from matplotlib import pyplot as plt
import math
import xlwt
import xlsxwriter
from collections import defaultdict

ini_term_dict = defaultdict(list)

excel_path = 'new.xlsx'

f = open(excel_path,'rb')
#     print(f)
df = pd.read_excel(f)



measure_length =  df['length'].tolist()
measure_time =  df['free_flow_time'].tolist()
ini_node = df['init_node'].tolist()
term_node = df['term_node'].tolist()
# print(len(set(ini_node+term_node)))
for i in range(len(ini_node)):
    # if ini_node[i] in ini_term_dict:
        ini_term_dict[ini_node[i]].append([term_node[i],measure_length[i]])
lst = []
ends = []
for i in ini_term_dict:
    lst.append(i)
    ends.append(ini_term_dict[i])

# print(lst)
# print(ends)



inf = 10e9

class graphMatrix:
    def __init__(self):
        self.dict={}  # 结点对应矩阵下标的字典
        self.length=0 # 矩阵的长度
        self.matrix=[]

    # 添加节点
    def addVertex(self,key):
        self.dict[key]=self.length # 键(顶点名称)->值(在矩阵中的下标)
        self.dict[self.length]=key # 键(在矩阵中的下标)->值(顶点名称)


        lst=[]
        for i in range(self.length):  # 创建一行长度为self.length,全是0的列表
            lst.append(inf)
        self.matrix.append(lst)

        for row in self.matrix:  # 为矩阵中每行添加新的一个单元，单元中值为0
            row.append(inf)

        self.length += 1

    # 添加无向边
    def addNoDirectLine(self,start,ends): #起始字符，终点(字符+权值)数组
        startIndex=self.dict[start] # 起点在二维矩阵中的下标
        for row in ends:  # 遍历终点数组
            endKey=row[0]
            endIndex=self.dict[endKey]  # 终点在二维矩阵中的下标
            weight=row[1]
            self.matrix[startIndex][endIndex]=weight  # 权值赋值
            self.matrix[endIndex][startIndex]=weight

    # 添加有向边
    def addDirectLine(self,start,ends):
        startIndex=self.dict[start] # 起点在二维矩阵中的下标
        for row in ends:  # 遍历终点数组
            endKey=row[0]
            endIndex=self.dict[endKey]  # 终点在二维矩阵中的下标
            weight=row[1]
            self.matrix[startIndex][endIndex]=weight  # 权值赋值

def construct_matrix():

    # lst=['北京','天津','郑州','青岛']
    # ends = [[['天津',138],['郑州',689]],[['郑州',700],['青岛',597]],[['青岛',725]]]
    gm=graphMatrix()
    for k in lst:
        gm.addVertex(k)
    for i in range(len(ends)):
        gm.addDirectLine(lst[i],ends[i])

    all_matrices = []
    print('\n')
    for i in range(len(gm.matrix)):
        all_matrices.append(gm.matrix[i])
        # print(gm.matrix[i])
    # print(all_matrices)
    return all_matrices


# construct_matrix()


######################dijistra求最短路径，边权重可以是长度和时间。

import numpy as np
def startwith(start: int, mgraph: list) -> list:
# 存储已知最小长度的节点编号 即是顺序
    passed = [start]
    nopass = [x for x in range(len(mgraph)) if x != start]

    dis = mgraph[start]
    # print(dis)
    # 创建字典 为直接与start节点相邻的节点初始化路径
    dict_ = {}
    dict_weight = {}
    for i in range(len(dis)):
        if dis[i] != np.inf:
            dict_[str(i)] = [start]
            dict_weight[str(i)] = [0]

    while len(nopass):
        idx = nopass[0]
        for i in nopass:
            if dis[i] < dis[idx]: 
                idx = i

        nopass.remove(idx)
        passed.append(idx)

        for i in nopass:
            if dis[idx] + mgraph[idx][i] < dis[i]:
                dis[i] = dis[idx] + mgraph[idx][i]
                dict_[str(i)] = dict_[str(idx)] + [idx] #如果之前经历过idx,而且当前到i的距离最短
                dict_weight[str(i)] =  dict_weight[str(idx)] + [dis[idx] + mgraph[idx][i]]

    # print(dict_)
    return dis,dict_,dict_weight

import pandas as pd


if __name__ == "__main__":
    # inf = 10086
    # mgraph = [[0, 1, 12, inf, inf, inf],
    #           [inf, 0, 9, 3, inf, inf],
    #           [inf, inf, 0, inf, 5, inf],
    #           [inf, inf, 4, 0, 13, 15],
    #           [inf, inf, inf ,inf, 0, 4],
    #           [inf, inf, inf, inf ,inf, 0]]
    matrix = construct_matrix()
 
    matrix = np.array(matrix)
    row, col = np.diag_indices_from(matrix)
    matrix[row,col] = 0
    matrix = matrix.tolist()
    # print(matrix)
    import csv

    # rows2 = ['abc1/ab1c','N']
    # for n in range(10):
    #     f = open("ok.csv", 'a',newline='')
    #     writer = csv.writer(f)
    #     writer.writerow(rows2)
    #     f.close()
    for j in range(416):
        start = j

        dis,dict_,dict_weight = startwith(start, matrix)#从0开始的path。
        # print(len(dis))#从start开始,到其他节点最短路径长度
        # print(len(dict_))#从start开始，到其他节点最短路径经过的路线
        # print(dict_)#从start开始，到其他节点最短路径经过的路线相应的距离
        # col_names = ['init_node',
        #         'min_length',
        #         'passed_node']
        init_node = []
        min_length = []
        passed_node = []
        f = open("paths_direct/"+str(j)+".csv", 'a',newline='')
        writer = csv.writer(f)
        for i in range(416):
            # init_node.append(i)
            
            
            print([dis[i]]+dict_[str(i)])
            writer.writerow([dis[i]]+dict_[str(i)])
        f.close()
            
        #     min_length.append(dis[i])
        #     passed_node.append(dict_[str(i)])
        # data = {
        # #   "init_node": init_node,
        # "min_length": min_length,
        # "passed_node": passed_node
        # }
        # df = pd.DataFrame(data)
        # print(df)
        # df.to_csv("paths/%s.csv",start)
'''

'''


import openpyxl
import numpy as np
workbook=openpyxl.load_workbook("new_Anaheim.xlsx")
worksheet=workbook.worksheets[0]
 
#在第一列之前插入一列
# worksheet.insert_cols(4)  #
 
# for index,col in enumerate(worksheet.rows):
#     print(index)
#     print(col)
#     if index==0:
#         continue
#     else:
#         col[4].value=col[2].value/col[3].value
# #枚举出来是tuple类型，从0开始计数
 
# workbook.save(filename="new_.xlsx")


# >>> a = 2
# >>> b = '{:08b}'.format(a)
ini_node = list(worksheet.columns)[0]
term_node = list(worksheet.columns)[1]
# print(ini_node.value)
# new_init_node = []
# new_term_node = []
record = {}
i = 0 
for cell in ini_node:
    # print(cell.value)
    if i != 0:
        v = int(cell.value)
        # print(type(v))
        if str(v)  in record:
            worksheet.cell(i+1,1,record[v])
        else:
            # bi =[ int(v) - 1 >>d & 1 for d in range(9)][::-1] #record[1] = [0,0,0,0,0,0,0,0,0]
            # record[v] = np.array(bi)
            # print(int(v) )
            record[v] = '{:09b}'.format(int(v) - 1)
            print(type(record[v]))
            worksheet.cell(i+1,1,record[v])
    i = i + 1
i = 0 
for cell in term_node:
    # print(cell.value)
    if i != 0:
        v = int(cell.value)
        if str(v)   in record:
            worksheet.cell(i+1,2,record[v])
        else:
            # bi =[ v - 1 >>d & 1 for d in range(9)][::-1] #record[1] = [0,0,0,0,0,0,0,0,0]
            record[v] = '{:09b}'.format(int(v) - 1)
            print(type(record[v]))
            
            worksheet.cell(i+1,2,record[v])
    i = i + 1

# print(worksheet)
# print(worksheet)

workbook.save(filename="new_.xlsx")

'''
# for i in range(len(ini_node)):
#     if ini_node[i] in record:
#         worksheet.cell(i+2,1,record[ini_node[i]])
#         new_init_node.append(record[ini_node[i]])
#     else:
#         bi =[ ini_node[i]-1 >>d & 1 for d in range(9)][::-1] #record[1] = [0,0,0,0,0,0,0,0,0]
#         record[ini_node[i]] = bi
#         new_init_node.append(record[ini_node[i]])

#     if term_node[i] in record:

#         new_term_node.append(record[term_node[i]])
#     else:
#         bi =[ term_node[i]-1 >>d & 1 for d in range(9)][::-1]
#         record[term_node[i]] = bi
#         new_term_node.append(record[term_node[i]])
    







# from mimetypes import init
# import pandas as pd
# path = '/Users/gaojunruo/Desktop/tomography/network_tomography-main/new.xlsx'
# f = open(path,'rb')
# #     print(f)
# df = pd.read_excel(f)
# init = df['init_node'].values
# end = df['term_node'].values
# label = []
# ini_term_pair = []
# for i in range(len(init)):
#     init_node = init[i]
#     end_node = end[i]
#     # ini_term_pair.append()
#     v = int(init_node)
#     init_list =[ v  >>d & 1 for d in range(9)][::-1] #record[1] = [0,0,0,0,0,0,0,0,0]
#     print('init_list')
#     print(init_list)
#     v = int(end_node)
#     term_list =[ v  >>d & 1 for d in range(9)][::-1] #record[1] = [0,0,0,0,0,0,0,0,0]
#     print('term_list')
#     print(term_list)
#     ini_term_pair.append(init_list+term_list)
# print('ini_term_pair')
# print(ini_term_pair)

'''

import sys
import os
import pandas as pd
from matplotlib import pyplot as plt
import math
import xlwt
import xlsxwriter
from collections import defaultdict
import openpyxl
 
import pandas as pd









ini_term_dict = defaultdict(list)

excel_path = 'new.xlsx'

f = open(excel_path,'rb')
#     print(f)
df = pd.read_excel(f)



measure_length =  df['length'].tolist()
measure_time =  df['free_flow_time'].tolist()
ini_node = df['init_node'].tolist()
term_node = df['term_node'].tolist()
# print(len(set(ini_node+term_node)))
for i in range(len(ini_node)):
    # if ini_node[i] in ini_term_dict:
        ini_term_dict[ini_node[i]].append([term_node[i],measure_length[i]])
lst = []
ends = []
for i in ini_term_dict:
    lst.append(i)
    ends.append(ini_term_dict[i])

# print(lst)
# print(ends)



inf = 10e9

class graphMatrix:
    def __init__(self):
        self.dict={}  # 结点对应矩阵下标的字典
        self.length=0 # 矩阵的长度
        self.matrix=[]

    # 添加节点
    def addVertex(self,key):
        self.dict[key]=self.length # 键(顶点名称)->值(在矩阵中的下标)
        self.dict[self.length]=key # 键(在矩阵中的下标)->值(顶点名称)


        lst=[]
        for i in range(self.length):  # 创建一行长度为self.length,全是0的列表
            lst.append(inf)
        self.matrix.append(lst)

        for row in self.matrix:  # 为矩阵中每行添加新的一个单元，单元中值为0
            row.append(inf)

        self.length += 1

    # 添加无向边
    def addNoDirectLine(self,start,ends): #起始字符，终点(字符+权值)数组
        startIndex=self.dict[start] # 起点在二维矩阵中的下标
        for row in ends:  # 遍历终点数组
            endKey=row[0]
            endIndex=self.dict[endKey]  # 终点在二维矩阵中的下标
            weight=row[1]
            self.matrix[startIndex][endIndex]=weight  # 权值赋值
            self.matrix[endIndex][startIndex]=weight

    # 添加有向边
    def addDirectLine(self,start,ends):
        startIndex=self.dict[start] # 起点在二维矩阵中的下标
        for row in ends:  # 遍历终点数组
            endKey=row[0]
            endIndex=self.dict[endKey]  # 终点在二维矩阵中的下标
            weight=row[1]
            self.matrix[startIndex][endIndex]=weight  # 权值赋值

def construct_matrix():

    # lst=['北京','天津','郑州','青岛']
    # ends = [[['天津',138],['郑州',689]],[['郑州',700],['青岛',597]],[['青岛',725]]]
    gm=graphMatrix()
    for k in lst:
        gm.addVertex(k)
    for i in range(len(ends)):
        gm.addDirectLine(lst[i],ends[i])

    all_matrices = []
    print('\n')
    for i in range(len(gm.matrix)):
        all_matrices.append(gm.matrix[i])
        # print(gm.matrix[i])
    # print(all_matrices)
    return all_matrices


# construct_matrix()


######################dijistra求最短路径，边权重可以是长度和时间。

import numpy as np
def startwith(start: int, mgraph: list) -> list:
# 存储已知最小长度的节点编号 即是顺序
    passed = [start]
    nopass = [x for x in range(len(mgraph)) if x != start]

    dis = mgraph[start]
    # print(dis)
    # 创建字典 为直接与start节点相邻的节点初始化路径
    dict_ = {}
    dict_weight = {}
    for i in range(len(dis)):
        if dis[i] != np.inf:
            dict_[str(i)] = [start]
            dict_weight[str(i)] = [0]

    while len(nopass):
        idx = nopass[0]
        for i in nopass:
            if dis[i] < dis[idx]: 
                idx = i

        nopass.remove(idx)
        passed.append(idx)

        for i in nopass:
            if dis[idx] + mgraph[idx][i] < dis[i]:
                dis[i] = dis[idx] + mgraph[idx][i]
                dict_[str(i)] = dict_[str(idx)] + [idx] #如果之前经历过idx,而且当前到i的距离最短
                dict_weight[str(i)] =  dict_weight[str(idx)] + [dis[idx] + mgraph[idx][i]]

    # print(dict_)
    return dis,dict_,dict_weight

import pandas as pd


if __name__ == "__main__":
    # inf = 10086
    # mgraph = [[0, 1, 12, inf, inf, inf],
    #           [inf, 0, 9, 3, inf, inf],
    #           [inf, inf, 0, inf, 5, inf],
    #           [inf, inf, 4, 0, 13, 15],
    #           [inf, inf, inf ,inf, 0, 4],
    #           [inf, inf, inf, inf ,inf, 0]]
    matrix = construct_matrix()
 
    matrix = np.array(matrix)
    row, col = np.diag_indices_from(matrix)
    matrix[row,col] = 0
    matrix = matrix.tolist()
    # print(matrix)
    import csv

    # rows2 = ['abc1/ab1c','N']
    # for n in range(10):
    #     f = open("ok.csv", 'a',newline='')
    #     writer = csv.writer(f)
    #     writer.writerow(rows2)
    #     f.close()
    term_node_list = []
    ini_node_list = []
    metric = []
    for j in range(416):
        start = j

        dis,dict_,dict_weight = startwith(start, matrix)#从0开始的path。
        ini_node = [j]*416
        term_node = list(range(416))#0-415
        ini_node_list = ini_node_list + ini_node
        term_node_list = term_node_list + term_node
        metric = metric + dis
    # print(metric)
        

        # print(type(dis))
        # print(len(dis))#从start开始,到其他节点最短路径长度
        # print(len(dict_))#从start开始，到其他节点最短路径经过的路线
        # print(dict_)#从start开始，到其他节点最短路径经过的路线相应的距离
        # col_names = ['init_node',
        #         'min_length',
        #         'passed_node']
        # init_node = []
        # min_length = []
        # passed_node = []
        # f = open("paths_direct/"+str(j)+".csv", 'a',newline='')
        # writer = csv.writer(f)
        # for i in range(416):
        #     # init_node.append(i)
            
            
        #     print([dis[i]]+dict_[str(i)])
        #     writer.writerow([dis[i]]+dict_[str(i)])
        # f.close()
            
        #     min_length.append(dis[i])
        #     passed_node.append(dict_[str(i)])
    print(type(ini_node_list))
    data = {
        'init_node': ini_node_list,
    'term_node': term_node_list,
    'metric': metric
    }
    df = pd.DataFrame(data)
    print(df['init_node'].values)
    # df['init_node'].values
    df.to_csv("augment_and_test.csv",index=False,sep=',')
'''

# import pandas as pd
# import random
# import numpy as np
# df_aug = pd.read_csv('augment_and_test.csv',names=['init_node','term_node','metric'],sep=',') 


# init_aug  = df_aug['init_node'].values
# end_aug  = df_aug['term_node'].values
# label_aug  = df_aug['metric'].values.tolist()
# # print(label_aug)
# list2=list(map(float,label_aug[1:]))  
# arr_var = np.var(list2)
# print(arr_var)
# arr_mean = np.mean(list2)
# print(arr_mean)

# # df = pd.read_csv('new.xlsx',header=None,sep=' ') 
# df = pd.read_excel('new.xlsx')
# init = df['init_node'].values
# end = df['term_node'].values
# label_list = df['length'].values.tolist()
# # print(label_list)
# arr_var = np.var(label_list)
# print(arr_var)
# arr_mean = np.mean(label_list)
# print(arr_mean)




# ini_end_label = list(zip(init_aug,end_aug,label_aug))
# slice = random.sample(ini_end_label, int(len(init_aug)*0.3) )
# print(slice)
# ini_term_pair = []
#     # label_list = []
# for i in range(len(init)):
#     init_node = init[i]
#     end_node = end[i]
#     # ini_term_pair.append()
#     v = int(init_node)
#     init_list =[ v  >>d & 1 for d in range(9)][::-1] #record[1] = [0,0,0,0,0,0,0,0,0]
#     # print(init_list)
#     v = int(end_node)
#     term_list =[ v  >>d & 1 for d in range(9)][::-1] #record[1] = [0,0,0,0,0,0,0,0,0]
#     ini_term_pair.append(init_list+term_list)
# for i in range(len(slice)):
#     init_node = slice[i][0]
#     v = int(init_node)
#     init_list =[ v  >>d & 1 for d in range(9)][::-1] #record[1] = [0,0,0,0,0,0,0,0,0]
#     end_node = slice[i][1]
#     v = int(end_node)
#     term_list =[ v  >>d & 1 for d in range(9)][::-1] #record[1] = [0,0,0,0,0,0,0,0,0]
#     label = slice[i][2]
#     label_list.append(int(float(label)))
#     ini_term_pair.append(init_list+term_list)
# print(ini_term_pair)
# print(label_list)




import numpy as np
import random

df_aug = pd.read_csv('augment_and_test.csv',names=['init_node','term_node','metric'],sep=',') 


init_aug  = df_aug['init_node'].values
end_aug  = df_aug['term_node'].values
label  = df_aug['metric'].values
label_aug = []
for i in range(len(label)):
    if i == 0:
        label_aug.append(label[i])
    else:
        label_aug.append(float(label[i])/1000)



# df = pd.read_csv('new.xlsx',header=None,sep=' ') 
df = pd.read_excel('new.xlsx')
init = df['init_node'].values
end = df['term_node'].values
label = df['length'].values.tolist()
label_list = []
for i in range(len(label)):
    # if i == 0:
    #     label_list.append(label[i])
    # else:
        label_list.append(float(label[i])/1000)

exit_node = set(init+end)


ini_end_label = list(zip(init_aug,end_aug,label_aug))
init_index = list(range(len(ini_end_label)))[1:]
slice = random.sample(init_index, int(len(init_index)*0.3) )
test_index = list(set(init_index)-set(slice))
# print(len(init_index))
# print(len(slice))
# print(len(test_index))
# init_aug
ini_term_pair = []
# label_list = []
# print(1)
for i in range(len(init)):
    init_node = init[i]
    end_node = end[i]
    # ini_term_pair.append()
    # print(type(init_node))
    # print(init_node)
    v = init_node
    # print(v)
    init_list =[ v  >>d & 1 for d in range(9)][::-1] #record[1] = [0,0,0,0,0,0,0,0,0]
    # print(init_list)
    v = end_node
    term_list =[ v  >>d & 1 for d in range(9)][::-1] #record[1] = [0,0,0,0,0,0,0,0,0]
    ini_term_pair.append(init_list+term_list)
# print(2)
#补充数据集
aug_data = []
# print(exit_node)
for i in range(len(ini_end_label)):
    # print(ini_end_label[i][0])
    if i == 0:
        continue
    init_node = int(ini_end_label[i][0])
    end_node = int(ini_end_label[i][1])
    if init_node in exit_node and end_node in exit_node:
        aug_data.append((init_node,end_node,ini_end_label[i][2]))
print(len(aug_data))
print(int(len(init_index)*0.3))

if len(aug_data) < int(len(init_index)*0.3) :
    


    for i in range(len(aug_data)):
        # if i != 0:
            init_node = aug_data[i][0]
            
            v = int(init_node)
            init_list =[ v  >>d & 1 for d in range(9)][::-1] #record[1] = [0,0,0,0,0,0,0,0,0]
            end_node = aug_data[i][1]
            v = int(end_node)
            term_list =[ v  >>d & 1 for d in range(9)][::-1] #record[1] = [0,0,0,0,0,0,0,0,0]
            label = aug_data[i][2]
            label_list.append(int(float(label)))
            ini_term_pair.append(init_list+term_list)
else:
    slice = random.sample(aug_data, int(len(init_index)*0.3) )
    for i in slice:
        # if i != 0:
            init_node = aug_data[i][0]
            
            v = int(init_node)
            init_list =[ v  >>d & 1 for d in range(9)][::-1] #record[1] = [0,0,0,0,0,0,0,0,0]
            end_node = aug_data[i][1]
            v = int(end_node)
            term_list =[ v  >>d & 1 for d in range(9)][::-1] #record[1] = [0,0,0,0,0,0,0,0,0]
            label = aug_data[i][2]
            label_list.append(int(float(label)))
            ini_term_pair.append(init_list+term_list)


test_ini_term_pair = []
test_label = []
# print(3)
#测试数据集
for i in init_index:
    if i not in slice and i != 0 :
        init_node = ini_end_label[i][0]
        # print(init_node)
        v = int(init_node)
        init_list =[ v  >>d & 1 for d in range(9)][::-1] #record[1] = [0,0,0,0,0,0,0,0,0]
        end_node = ini_end_label[i][1]
        v = int(end_node)
        term_list =[ v  >>d & 1 for d in range(9)][::-1] #record[1] = [0,0,0,0,0,0,0,0,0]
        label = ini_end_label[i][2]
        test_label.append(int(float(label)))
        test_ini_term_pair.append(init_list+term_list)
# print(4)
# ini_term_pair, label_list,test_ini_term_pair,test_label


ini_term_pair=np.array(ini_term_pair)
label_list=np.array(label_list)
test_ini_term_pair=np.array(test_ini_term_pair)
test_label=np.array(test_label)
print(ini_term_pair)
print(label_list)
print(test_ini_term_pair)
print(test_label)
np.save('ini_term_pair.npy',ini_term_pair) # 保存为.npy格式
np.save('label_list.npy',label_list) # 保存为.npy格式
np.save('test_ini_term_pair.npy',test_ini_term_pair) # 保存为.npy格式
np.save('test_label.npy',test_label) # 保存为.npy格式